"""Импорт проектов и дедлайнов из CSV/Excel. / Import projects & deadlines from CSV/Excel.

Формат — «длинная» таблица, одна строка = один дедлайн этапа одного проекта:

    project          | stage | planned_date | completed
    -----------------|-------|--------------|----------
    Офис А           | 6     | 2026-06-25   | no
    Офис А           | 5     | 2026-06-20   | yes
    Кафе Б           | 14    | 2026-07-02   |

Long format, one row = one stage deadline of one project. Projects are created
if missing, otherwise the named stages are updated. Stages not listed are left
untouched. Column headers are matched case-insensitively in RU and EN.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime

from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import Project, ensure_stage_rows
from .stages import STAGE_COUNT

# Синонимы заголовков колонок. / Accepted header synonyms.
_HEADERS = {
    "project": {"project", "проект", "объект", "name", "название"},
    "stage": {"stage", "stage_index", "этап", "stage_no", "№", "no", "номер этапа"},
    "planned_date": {
        "planned_date",
        "date",
        "дата",
        "дедлайн",
        "deadline",
        "срок",
        "плановая дата",
    },
    "completed": {"completed", "done", "выполнено", "готово", "completed?", "status"},
}
_TRUE = {"1", "yes", "y", "true", "да", "готово", "done", "x", "✓", "+"}


class ImportError_(Exception):
    pass


def _norm_header(name: str) -> str | None:
    key = (name or "").strip().lower()
    for canonical, syns in _HEADERS.items():
        if key in syns:
            return canonical
    return None


def parse_upload(filename: str, content: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content)
    if name.endswith(".csv") or name.endswith(".txt"):
        return _parse_csv(content)
    # по содержимому: попробуем как xlsx (ZIP-сигнатура), иначе CSV
    if content[:2] == b"PK":
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return []
    header_map = {i: _norm_header(h) for i, h in enumerate(rows[0])}
    if "project" not in header_map.values() or "stage" not in header_map.values():
        raise ImportError_(
            "Не найдены колонки 'project' и 'stage'. / Columns 'project' and 'stage' are required."
        )
    out: list[dict] = []
    for raw in rows[1:]:
        rec = {}
        for i, value in enumerate(raw):
            col = header_map.get(i)
            if col:
                rec[col] = value
        if rec.get("project"):
            out.append(rec)
    return out


def _parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_map = {i: _norm_header(str(h) if h is not None else "") for i, h in enumerate(rows[0])}
    if "project" not in header_map.values() or "stage" not in header_map.values():
        raise ImportError_(
            "Не найдены колонки 'project' и 'stage'. / Columns 'project' and 'stage' are required."
        )
    out: list[dict] = []
    for raw in rows[1:]:
        rec = {}
        for i, value in enumerate(raw):
            col = header_map.get(i)
            if col is not None:
                rec[col] = value
        if rec.get("project") not in (None, ""):
            out.append(rec)
    return out


def _to_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ImportError_(f"Не понял дату: '{text}' / Unrecognized date: '{text}'")


def _to_stage(value) -> int:
    text = str(value).strip()
    # допускаем "6" и "6. Возведение..." / accept "6" and "6. ..."
    head = text.split(".", 1)[0].split()[0] if text else ""
    if not head.isdigit():
        raise ImportError_(
            f"Этап должен быть числом 1..{STAGE_COUNT}: '{text}' / "
            f"Stage must be a number 1..{STAGE_COUNT}: '{text}'"
        )
    n = int(head)
    if not 1 <= n <= STAGE_COUNT:
        raise ImportError_(f"Этап вне диапазона 1..{STAGE_COUNT}: {n}")
    return n


def _to_bool(value) -> bool:
    if value in (None, ""):
        return False
    return str(value).strip().lower() in _TRUE


def apply_rows(db: Session, rows: list[dict]) -> dict:
    """Применить строки к БД. / Apply parsed rows. Returns a summary dict."""
    cache: dict[str, Project] = {}
    created: set[str] = set()
    updated: set[str] = set()
    applied = 0
    errors: list[str] = []

    for line_no, rec in enumerate(rows, start=2):  # 1 = header
        name = str(rec.get("project", "")).strip()
        if not name:
            continue
        try:
            stage_index = _to_stage(rec.get("stage"))
            planned = _to_date(rec.get("planned_date"))
            completed = _to_bool(rec.get("completed"))
        except ImportError_ as exc:
            errors.append(f"стр. {line_no} / row {line_no}: {exc}")
            continue

        project = cache.get(name)
        if project is None:
            project = db.scalar(select(Project).where(Project.name == name))
            if project is None:
                project = Project(name=name)
                ensure_stage_rows(project)
                db.add(project)
                db.flush()
                created.add(name)
            else:
                ensure_stage_rows(project)
                updated.add(name)
            cache[name] = project

        row = project.stage_map[stage_index]
        if planned is not None:
            row.planned_date = planned
        if "completed" in rec:
            row.completed = completed
            row.completed_at = datetime.utcnow() if completed else None
        applied += 1

    db.commit()
    return {
        "applied": applied,
        "created": sorted(created),
        "updated": sorted(updated - created),
        "errors": errors,
        "total_rows": len(rows),
    }


TEMPLATE_CSV = (
    "project,stage,planned_date,completed\n"
    "Офис А / Office A,5,2026-06-20,yes\n"
    "Офис А / Office A,6,2026-06-25,no\n"
    "Кафе Б / Cafe B,14,2026-07-02,\n"
)
